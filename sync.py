import os
import json
import requests
import datetime
from zoneinfo import ZoneInfo
from openpyxl import load_workbook
from io import BytesIO

# ==============================
# DROPBOX AUTH
# ==============================

REFRESH_TOKEN = os.environ['DP_TOKEN']
APP_KEY = os.environ['DP_APP_TOKEN']
APP_SECRET = os.environ['DP_SECRET']

def get_access_token():
    r = requests.post(
        'https://api.dropbox.com/oauth2/token',
        data={
            'grant_type': 'refresh_token',
            'refresh_token': REFRESH_TOKEN,
            'client_id': APP_KEY,
            'client_secret': APP_SECRET,
        }
    )
    print('Token svar:', r.status_code)
    r.raise_for_status()
    return r.json()['access_token']

ACCESS_TOKEN = get_access_token()
HEADERS = {
    'Authorization': f'Bearer {ACCESS_TOKEN}',
    'Content-Type': 'application/json'
}

DROPBOX_FOLDER = ''
OUTPUT_DIR = 'data'

FILE_MAP = {
    'mål': 'mal',
    'mal': 'mal',
    'produktionsr': 'utfall',
    'uträkning': 'utfall',
    'utfall': 'utfall',
}

# Rådata-blad som ska exporteras separat
RAWDATA_SHEETS = {
    '0.1 data försäljning': 'rawdata_orders',
    '0.0 data timmar': 'rawdata_timmar',
}

os.makedirs(OUTPUT_DIR, exist_ok=True)

# ==============================
# DROPBOX FILE FUNCTIONS
# ==============================

def list_files():
    all_entries = []
    r = requests.post(
        'https://api.dropboxapi.com/2/files/list_folder',
        headers=HEADERS,
        json={'path': DROPBOX_FOLDER, 'recursive': True}
    )
    print('list_folder status:', r.status_code)
    r.raise_for_status()
    data = r.json()
    all_entries.extend(data.get('entries', []))

    while data.get('has_more'):
        r = requests.post(
            'https://api.dropboxapi.com/2/files/list_folder/continue',
            headers=HEADERS,
            json={'cursor': data['cursor']}
        )
        r.raise_for_status()
        data = r.json()
        all_entries.extend(data.get('entries', []))

    return all_entries

def download_file(path):
    r = requests.post(
        'https://content.dropboxapi.com/2/files/download',
        headers={
            'Authorization': f'Bearer {ACCESS_TOKEN}',
            'Dropbox-API-Arg': json.dumps({'path': path})
        }
    )
    r.raise_for_status()
    return r.content

# ==============================
# ROBUST EXCEL PARSER
# ==============================

def excel_to_json(content):
    wb = load_workbook(BytesIO(content), data_only=True)
    result = {}
    rawdata = {}

    def norm(x):
        return str(x).strip().lower().replace(':','') if x is not None else ''

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Check if this is a rawdata sheet
        sheet_lower = sheet_name.strip().lower()
        raw_key = None
        for pattern, key in RAWDATA_SHEETS.items():
            if sheet_lower.startswith(pattern) or pattern in sheet_lower:
                raw_key = key
                break

        if raw_key:
            # Parse rawdata sheet as flat array of rows
            raw_rows = parse_rawdata_sheet(sheet_name, rows)
            if raw_rows:
                rawdata[raw_key] = rawdata.get(raw_key, []) + raw_rows
                print(f'✅ RÅDATA {sheet_name}: {len(raw_rows)} rader')
            else:
                print(f'⚠️ RÅDATA {sheet_name}: inga rader efter parsing')
            continue

        # Skip other "0." sheets that aren't mapped
        if sheet_name.strip().startswith('0'):
            print(f'Hoppar över okänt rådata-blad: {sheet_name}')
            continue

        # === Normal sheet parsing (unchanged) ===
        header_row_idx = None
        for i in range(min(60, len(rows))):
            if any(norm(c) in ('datum', 'date') for c in rows[i]):
                header_row_idx = i
                break

        if header_row_idx is None:
            print(f'⚠️ Ingen header med DATUM hittades i: {sheet_name}')
            continue

        headers = [str(h).strip() if h is not None else '' for h in rows[header_row_idx]]

        date_col = None
        for h in headers:
            if norm(h) in ('datum', 'date'):
                date_col = h
                break

        if not date_col:
            print(f'⚠️ Hittade header men ingen datum-kolumn i: {sheet_name}')
            continue

        sheet_data = {}

        for row in rows[header_row_idx + 1:]:
            row_dict = dict(zip(headers, row))
            date_val = row_dict.get(date_col)

            if not date_val:
                continue

            date_key = parse_date(date_val)
            if not date_key:
                continue

            clean = clean_row(row_dict)
            if clean:
                sheet_data[date_key] = clean

        if sheet_data:
            result[sheet_name] = sheet_data
            print(f'✅ {sheet_name}: {len(sheet_data)} rader')
        else:
            print(f'⚠️ {sheet_name}: inga rader efter parsing')

    return result, rawdata


def parse_rawdata_sheet(sheet_name, rows):
    """Parse a rawdata sheet into a flat list of row dicts."""
    def norm(x):
        return str(x).strip().lower().replace(':','') if x is not None else ''

    # Find header row
    header_row_idx = None
    for i in range(min(20, len(rows))):
        cells = [norm(c) for c in rows[i]]
        if 'datum' in cells or 'date' in cells:
            header_row_idx = i
            break
        # Also check for 'säljare' or 'projekt' as header indicators
        if 'säljare' in cells and 'projekt' in cells:
            header_row_idx = i
            break

    if header_row_idx is None:
        print(f'⚠️ Ingen header hittades i rådata: {sheet_name}')
        return []

    headers = [str(h).strip() if h is not None else '' for h in rows[header_row_idx]]

    parsed = []
    for row in rows[header_row_idx + 1:]:
        row_dict = dict(zip(headers, row))

        # Skip completely empty rows
        if all(v is None or str(v).strip() == '' for v in row):
            continue

        clean = clean_row(row_dict)
        if not clean:
            continue

        # Convert any date fields to string
        for k, v in clean.items():
            if isinstance(v, (datetime.datetime, datetime.date)):
                clean[k] = v.strftime('%Y-%m-%d')

        # Ensure Datum is a proper date string
        date_val = clean.get('Datum') or clean.get('datum')
        if date_val:
            parsed_date = parse_date(date_val)
            if parsed_date:
                clean['Datum'] = parsed_date

        parsed.append(clean)

    return parsed


def parse_date(date_val):
    """Convert various date formats to YYYY-MM-DD string."""
    if isinstance(date_val, datetime.datetime):
        return date_val.strftime('%Y-%m-%d')
    elif isinstance(date_val, datetime.date):
        return date_val.strftime('%Y-%m-%d')
    elif isinstance(date_val, (int, float)):
        base = datetime.datetime(1899, 12, 30)
        try:
            return (base + datetime.timedelta(days=float(date_val))).strftime('%Y-%m-%d')
        except Exception:
            return None
    elif isinstance(date_val, str):
        s = date_val.strip()
        try:
            if '-' in s:
                return s[:10]
            elif '/' in s:
                parts = s.split('/')
                if len(parts[0]) == 4:
                    return f"{parts[0]}-{parts[1].zfill(2)}-{parts[2].zfill(2)}"
                else:
                    return f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
        except Exception:
            return None
    return None


def clean_row(row_dict):
    """Clean a row dict, removing empty keys and converting types."""
    clean = {}
    for k, v in row_dict.items():
        if not k or str(k).strip() == '':
            continue
        if isinstance(v, datetime.datetime):
            clean[k] = v.strftime('%Y-%m-%d')
        elif isinstance(v, datetime.date):
            clean[k] = v.strftime('%Y-%m-%d')
        elif isinstance(v, (int, float)):
            clean[k] = v
        elif v is not None and str(v).strip() != '':
            clean[k] = str(v)
    return clean


# ==============================
# FILE TYPE DETECTION
# ==============================

def detect_type(filename):
    fn = filename.lower()
    for keyword, typ in FILE_MAP.items():
        if keyword in fn:
            return typ
    return None

# ==============================
# MAIN SYNC LOOP
# ==============================

files = list_files()
print(f'Hittade {len(files)} filer/mappar i Dropbox (rekursivt)')

all_data = {'utfall': {}, 'mal': {}}
all_rawdata = {'rawdata_orders': [], 'rawdata_timmar': []}

for f in files:
    if f['.tag'] != 'file':
        continue

    name = f['name']
    if not name.endswith(('.xlsx', '.xls')):
        continue

    typ = detect_type(name)
    if not typ:
        print(f'Okänd filtyp: {name}, hoppar över')
        continue

    print(f'Laddar ner: {name} ({typ}) från {f["path_lower"]}')
    content = download_file(f['path_lower'])
    data, rawdata = excel_to_json(content)

    # Merge normal sheet data
    for sheet_name, sheet_data in data.items():
        if sheet_name not in all_data[typ]:
            all_data[typ][sheet_name] = {}
        all_data[typ][sheet_name].update(sheet_data)

    # Merge rawdata
    for raw_key, raw_rows in rawdata.items():
        all_rawdata[raw_key].extend(raw_rows)

    print(f'  → {len(data)} flikar, {sum(len(v) for v in rawdata.values())} rådata-rader från {name}')

# Save normal JSON files
for typ, data in all_data.items():
    if data:
        out_path = os.path.join(OUTPUT_DIR, f'{typ}.json')
        with open(out_path, 'w', encoding='utf-8') as fh:
            json.dump(data, fh, ensure_ascii=False, indent=2)
        print(f'Sparade {out_path} med {len(data)} flikar')

# Save rawdata JSON files
for raw_key, raw_rows in all_rawdata.items():
    if raw_rows:
        out_path = os.path.join(OUTPUT_DIR, f'{raw_key}.json')
        with open(out_path, 'w', encoding='utf-8') as fh:
            json.dump(raw_rows, fh, ensure_ascii=False, indent=2)
        print(f'Sparade {out_path} med {len(raw_rows)} rader')

# ==============================
# SAVE LAST SYNC TIME
# ==============================

ts = datetime.datetime.now(ZoneInfo('Europe/Stockholm')).strftime('%Y-%m-%d %H:%M')
with open(os.path.join(OUTPUT_DIR, 'last_synced.json'), 'w') as fh:
    json.dump({'last_synced': ts}, fh)

print(f'Klar! Synkad {ts}')
